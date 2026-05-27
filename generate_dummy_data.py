"""
Generates dummy cc_transactions.xlsx for demo/showcase purposes.
Run once: python generate_dummy_data.py
Overwrites data/cc_transactions.xlsx.
"""
import hashlib
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

EXCEL_PATH = Path("data/cc_transactions.xlsx")
BANK_ORDER = ["BDO", "BPI_Lincoln", "BPI_Xherine"]

def md5_hash(bank, dt, tx, amt):
    raw = f"{bank}{dt}{tx}{amt:.2f}"
    return hashlib.md5(raw.encode()).hexdigest()

# Each row: (statement_date, date, transaction, amount)
BDO_ROWS = [
    # Cycle 2026-01  (stmt 2026-02-01)
    ("2026-02-01", "2026-01-16", "PETRON COMMONWEALTH",          3200.00),
    ("2026-02-01", "2026-01-18", "GRAB RIDE",                     180.00),
    ("2026-02-01", "2026-01-20", "SMC SKYWAY TOLL",                220.00),
    ("2026-02-01", "2026-01-22", "GRAB FOOD",                     380.00),
    ("2026-02-01", "2026-01-24", "MERCURY DRUG COMMONWEALTH",     450.00),
    ("2026-02-01", "2026-01-25", "SMC SKYWAY TOLL",                220.00),
    ("2026-02-01", "2026-01-27", "JOLLIBEE DRIVE THRU",           520.00),
    ("2026-02-01", "2026-01-28", "PETRON BALINTAWAK",            2800.00),
    ("2026-02-01", "2026-01-29", "GRAB FOOD",                     340.00),
    ("2026-02-01", "2026-01-30", "SM SUPERMARKET NORTH",         2100.00),
    ("2026-02-01", "2026-02-01", "FINANCE CHARGE",               1050.00),
    # Cycle 2026-02  (stmt 2026-03-01)
    ("2026-03-01", "2026-02-16", "PETRON COMMONWEALTH",          3400.00),
    ("2026-03-01", "2026-02-17", "GRAB RIDE",                     200.00),
    ("2026-03-01", "2026-02-18", "SMC SKYWAY TOLL",                220.00),
    ("2026-03-01", "2026-02-20", "GRAB FOOD",                     420.00),
    ("2026-03-01", "2026-02-22", "MERCURY DRUG COMMONWEALTH",     680.00),
    ("2026-03-01", "2026-02-23", "SMC SKYWAY TOLL",                220.00),
    ("2026-03-01", "2026-02-24", "MCDONALDS DRIVE THRU",          380.00),
    ("2026-03-01", "2026-02-26", "GRAB FOOD",                     460.00),
    ("2026-03-01", "2026-02-27", "PETRON QUEZON AVE",            2950.00),
    ("2026-03-01", "2026-02-28", "GREENWICH PIZZA TRINOMA",       580.00),
    ("2026-03-01", "2026-03-01", "FINANCE CHARGE",               1120.00),
    # Cycle 2026-03  (stmt 2026-04-01)
    ("2026-04-01", "2026-03-16", "PETRON COMMONWEALTH",          3100.00),
    ("2026-04-01", "2026-03-17", "GRAB RIDE",                     220.00),
    ("2026-04-01", "2026-03-18", "SMC SKYWAY TOLL",                220.00),
    ("2026-04-01", "2026-03-20", "GRAB FOOD",                     490.00),
    ("2026-04-01", "2026-03-22", "MERCURY DRUG COMMONWEALTH",     550.00),
    ("2026-04-01", "2026-03-24", "SMC SKYWAY TOLL",                220.00),
    ("2026-04-01", "2026-03-25", "GRAB FOOD",                     380.00),
    ("2026-04-01", "2026-03-27", "PETRON BALINTAWAK",            3300.00),
    ("2026-04-01", "2026-03-28", "YELLOW CAB PIZZA BGC",          890.00),
    ("2026-04-01", "2026-03-30", "WATSONS TRINOMA",               620.00),
    ("2026-04-01", "2026-04-01", "FINANCE CHARGE",                980.00),
    # Cycle 2026-04  (stmt 2026-05-01)
    ("2026-05-01", "2026-04-16", "PETRON COMMONWEALTH",          3250.00),
    ("2026-05-01", "2026-04-17", "GRAB RIDE",                     190.00),
    ("2026-05-01", "2026-04-18", "SMC SKYWAY TOLL",                220.00),
    ("2026-05-01", "2026-04-20", "GRAB FOOD",                     450.00),
    ("2026-05-01", "2026-04-22", "MERCURY DRUG COMMONWEALTH",     720.00),
    ("2026-05-01", "2026-04-23", "SMC SKYWAY TOLL",                220.00),
    ("2026-05-01", "2026-04-25", "GRAB FOOD",                     390.00),
    ("2026-05-01", "2026-04-26", "PETRON QUEZON AVE",            2900.00),
    ("2026-05-01", "2026-04-28", "STARBUCKS BGC",                 580.00),
    ("2026-05-01", "2026-04-29", "WATSONS AYALA",                 490.00),
    ("2026-05-01", "2026-05-01", "FINANCE CHARGE",               1050.00),
]

BPI_LINCOLN_ROWS = [
    # Cycle 2026-01  (stmt 2026-01-25)
    ("2026-01-25", "2026-01-15", "GRAB FOOD",                     350.00),
    ("2026-01-25", "2026-01-16", "LAZADA PHILIPPINES",           2890.00),
    ("2026-01-25", "2026-01-17", "SHOPWISE ALABANG",             4350.00),
    ("2026-01-25", "2026-01-18", "APPLE ITUNES STORE",            349.00),
    ("2026-01-25", "2026-01-19", "GRAB RIDE",                     240.00),
    ("2026-01-25", "2026-01-20", "MERCURY DRUG ALABANG",          580.00),
    ("2026-01-25", "2026-01-21", "SPOTIFY SUBSCRIPTION",          169.00),
    ("2026-01-25", "2026-01-21", "TIKTOK SHOP",                  1450.00),
    ("2026-01-25", "2026-01-22", "GRAB FOOD",                     410.00),
    ("2026-01-25", "2026-01-23", "NATIONAL BOOKSTORE ALABANG",    890.00),
    ("2026-01-25", "2026-01-23", "WATSONS ALABANG",               680.00),
    ("2026-01-25", "2026-01-24", "WALTER MART ALABANG",          3120.00),
    # Cycle 2026-02  (stmt 2026-02-25)
    ("2026-02-25", "2026-02-15", "GRAB FOOD",                     420.00),
    ("2026-02-25", "2026-02-16", "WALTER MART ALABANG",          3680.00),
    ("2026-02-25", "2026-02-17", "LAZADA PHILIPPINES",           1890.00),
    ("2026-02-25", "2026-02-18", "APPLE ITUNES STORE",            349.00),
    ("2026-02-25", "2026-02-19", "GRAB RIDE",                     340.00),
    ("2026-02-25", "2026-02-20", "MERCURY DRUG ALABANG",          480.00),
    ("2026-02-25", "2026-02-21", "TIKTOK SHOP",                  2100.00),
    ("2026-02-25", "2026-02-22", "GRAB FOOD",                     560.00),
    ("2026-02-25", "2026-02-23", "TIMEZONE GATEWAY",             1450.00),
    ("2026-02-25", "2026-02-24", "SPOTIFY SUBSCRIPTION",          169.00),
    ("2026-02-25", "2026-02-24", "WATSONS ALABANG",               750.00),
    ("2026-02-25", "2026-02-25", "SHOPWISE ALABANG",             4230.00),
    ("2026-02-25", "2026-02-25", "FINANCE CHARGE",                920.00),
    # Cycle 2026-03  (stmt 2026-03-25)
    ("2026-03-25", "2026-03-15", "GRAB FOOD",                     390.00),
    ("2026-03-25", "2026-03-16", "LAZADA PHILIPPINES",           3200.00),
    ("2026-03-25", "2026-03-17", "SHOPWISE ALABANG",             3890.00),
    ("2026-03-25", "2026-03-18", "APPLE ITUNES STORE",            349.00),
    ("2026-03-25", "2026-03-19", "GRAB RIDE",                     290.00),
    ("2026-03-25", "2026-03-20", "MERCURY DRUG ALABANG",          830.00),
    ("2026-03-25", "2026-03-21", "SPOTIFY SUBSCRIPTION",          169.00),
    ("2026-03-25", "2026-03-22", "TIKTOK SHOP",                  1650.00),
    ("2026-03-25", "2026-03-23", "GRAB FOOD",                     480.00),
    ("2026-03-25", "2026-03-24", "BOOKSALE INC ALABANG",          650.00),
    ("2026-03-25", "2026-03-24", "WATSONS ALABANG",               920.00),
    ("2026-03-25", "2026-03-25", "WALTER MART ALABANG",          4560.00),
    ("2026-03-25", "2026-03-25", "MAXIMS HOTEL MANILA",          8500.00),
    ("2026-03-25", "2026-03-25", "FINANCE CHARGE",               1100.00),
    # Cycle 2026-04  (stmt 2026-04-26)
    ("2026-04-26", "2026-04-15", "GRAB FOOD",                     410.00),
    ("2026-04-26", "2026-04-16", "LAZADA PHILIPPINES",           2350.00),
    ("2026-04-26", "2026-04-17", "SHOPWISE ALABANG",             4100.00),
    ("2026-04-26", "2026-04-18", "APPLE ITUNES STORE",            349.00),
    ("2026-04-26", "2026-04-19", "GRAB RIDE",                     310.00),
    ("2026-04-26", "2026-04-20", "MERCURY DRUG ALABANG",          560.00),
    ("2026-04-26", "2026-04-21", "SPOTIFY SUBSCRIPTION",          169.00),
    ("2026-04-26", "2026-04-22", "GRAB FOOD",                     490.00),
    ("2026-04-26", "2026-04-23", "TIKTOK SHOP",                  1780.00),
    ("2026-04-26", "2026-04-24", "NATIONAL BOOKSTORE ALABANG",    890.00),
    ("2026-04-26", "2026-04-25", "WATSONS ALABANG",               680.00),
    ("2026-04-26", "2026-04-26", "FINANCE CHARGE",                980.00),
]

HEADER_FILL = PatternFill("solid", fgColor="1E293B")
HEADER_FONT = Font(bold=True, color="94A3B8", size=10)

def write_sheet(ws, bank, rows):
    ws.append(["Statement Date", "Date", "Transaction", "Amount", "Hash"])
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left")

    for stmt_dt, tx_dt, tx_name, amt in rows:
        h = md5_hash(bank, tx_dt, tx_name, amt)
        ws.append([stmt_dt, tx_dt, tx_name, amt, h])

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 36


def main():
    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for bank in BANK_ORDER:
        ws = wb.create_sheet(bank)
        if bank == "BDO":
            write_sheet(ws, bank, BDO_ROWS)
        elif bank == "BPI_Lincoln":
            write_sheet(ws, bank, BPI_LINCOLN_ROWS)
        else:
            ws.append(["Statement Date", "Date", "Transaction", "Amount", "Hash"])

    wb.save(EXCEL_PATH)
    print(f"[OK] Dummy data written -> {EXCEL_PATH}")
    bdo_count = len(BDO_ROWS)
    bpi_count = len(BPI_LINCOLN_ROWS)
    print(f"     BDO:         {bdo_count} transactions")
    print(f"     BPI_Lincoln: {bpi_count} transactions")
    print(f"     Total:       {bdo_count + bpi_count} transactions")


if __name__ == "__main__":
    main()
