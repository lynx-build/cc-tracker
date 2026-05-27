import json
import subprocess
import sys
from datetime import datetime, date
from pathlib import Path
from collections import defaultdict

import openpyxl

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ── Paths ─────────────────────────────────────────────────────────────────────
EXCEL_PATH     = Path("data/cc_transactions.xlsx")
DASHBOARD_PATH = Path("dashboard/dashboard.html")
DOCS_PATH      = Path("docs/index.html")
LOG_PATH       = Path("logs/run_log.txt")

# ── Bank config ───────────────────────────────────────────────────────────────
BANK_ORDER = ["BDO", "BPI_Lincoln", "BPI_Xherine"]

BANK_COLORS = {
    "BDO":         "#1a56db",
    "BPI_Lincoln": "#e02424",
    "BPI_Xherine": "#9ca3af",
}
DEFAULT_BANK_COLOR = "#6b7280"

# ── Cycle config ──────────────────────────────────────────────────────────────
START_CYCLE = (2026, 1)   # Jan 15, 2026

# ── Category groups ───────────────────────────────────────────────────────────
MERCHANT_GROUPS = {
    "Finance Charge":  {"keywords": ["finance charge"],                              "icon": "💳"},
    "Grab":            {"keywords": ["grab"],                                        "icon": "🚗"},
    "Online Shopping": {"keywords": ["lazada", "tiktok"],                            "icon": "🛍️"},
    "Subscriptions":   {"keywords": ["apple", "itunes", "app store", "spotify"],     "icon": "📱"},
    "Pharmacy":        {"keywords": ["mercurydrug", "watsons"],                      "icon": "💊"},
    "Toll fee":        {"keywords": ["smc skyway"],                                  "icon": "🛣️"},
    "Gas":             {"keywords": ["petron", " gas"],                              "icon": "⛽"},
    "Groceries":       {"keywords": ["shopwise", "wms", "waltermart", "wmart"],      "icon": "🛒"},
}

CATEGORY_COLORS = {
    "Finance Charge":  "#ef4444",
    "Grab":            "#06b6d4",
    "Online Shopping": "#f59e0b",
    "Subscriptions":   "#8b5cf6",
    "Pharmacy":        "#ec4899",
    "Toll fee":        "#f97316",
    "Gas":             "#eab308",
    "Groceries":       "#10b981",
    "Other":           "#475569",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def bank_display_parts(key: str) -> list:
    parts = key.split("_", 1)
    return [parts[0], parts[1] if len(parts) > 1 else ""]


def get_last_updated() -> str:
    try:
        text  = LOG_PATH.read_text(encoding="utf-8")
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        if lines:
            return lines[-1].split(" | ")[0]
    except Exception:
        pass
    return datetime.now().strftime("%Y-%m-%d %H:%M")


def _matches_group(name: str, keywords: list) -> bool:
    n = name.lower()
    return any(kw in n for kw in keywords)


def _get_category(tx_name: str) -> str:
    for gname, gcfg in MERCHANT_GROUPS.items():
        if _matches_group(tx_name, gcfg["keywords"]):
            return gname
    return "Other"


def get_cycle_key(statement_date_str: str):
    try:
        d = datetime.strptime(statement_date_str, "%Y-%m-%d")
    except (ValueError, TypeError):
        return None
    if d.day <= 5:
        if d.month == 1:
            return (d.year - 1, 12)
        return (d.year, d.month - 1)
    return (d.year, d.month)


def cycle_str(ck: tuple) -> str:
    return f"{ck[0]:04d}-{ck[1]:02d}"


def cycle_label(ck_str: str) -> str:
    MN = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    y, m = int(ck_str[:4]), int(ck_str[5:7])
    nm = 1 if m == 12 else m + 1
    ny = y + 1 if m == 12 else y
    return f"{MN[m-1]} 15 – {MN[nm-1]} 5 '{str(ny)[2:]}"


def fmt_stmt_date(d_str: str) -> str:
    try:
        dt = datetime.strptime(d_str, "%Y-%m-%d")
        return dt.strftime("%b %d, %Y").replace(" 0", " ")
    except Exception:
        return d_str


# ── Data loading ──────────────────────────────────────────────────────────────

def load_transactions() -> dict:
    result = {b: [] for b in BANK_ORDER}

    if not EXCEL_PATH.exists():
        print("[WARN] cc_transactions.xlsx not found — dashboard will show empty data")
        return result

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    for bank in BANK_ORDER:
        if bank not in wb.sheetnames:
            continue
        ws = wb[bank]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4 or not row[1]:
                continue

            def _date_str(v):
                if isinstance(v, (datetime, date)):
                    return v.strftime("%Y-%m-%d")
                return str(v).strip() if v else ""

            result[bank].append({
                "statement_date": _date_str(row[0]),
                "date":           _date_str(row[1]),
                "transaction":    str(row[2] or "").strip(),
                "amount":         float(row[3]) if row[3] is not None else 0.0,
                "bank":           bank,
            })

    wb.close()
    return result


# ── Summary computation ───────────────────────────────────────────────────────

def compute_summaries(data: dict) -> dict:
    all_cycles_set    = set()
    cycle_total       = defaultdict(float)
    cycle_bank_totals = defaultdict(lambda: defaultdict(float))
    cycle_bank_counts = defaultdict(lambda: defaultdict(int))
    cycle_bank_stmts  = defaultdict(lambda: defaultdict(str))
    cycle_cat_totals  = defaultdict(lambda: defaultdict(float))
    cycle_cat_counts  = defaultdict(lambda: defaultdict(int))
    cycle_fc          = defaultdict(float)
    cycle_merch       = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    cycle_spend       = {b: defaultdict(float) for b in BANK_ORDER}
    all_tx            = []
    stmt_dates_set    = set()

    for bank, rows in data.items():
        for row in rows:
            sdt = row["statement_date"]
            ck  = get_cycle_key(sdt)
            if ck is None or ck < START_CYCLE:
                continue

            ck_s = cycle_str(ck)
            amt  = row["amount"]
            tx   = row["transaction"]
            cat  = _get_category(tx)

            all_cycles_set.add(ck_s)
            cycle_total[ck_s]             += amt
            cycle_bank_totals[ck_s][bank] += amt
            cycle_bank_counts[ck_s][bank] += 1
            cycle_spend[bank][ck_s]       += amt

            if sdt and sdt > cycle_bank_stmts[ck_s][bank]:
                cycle_bank_stmts[ck_s][bank] = sdt
            if sdt:
                stmt_dates_set.add(sdt)

            cycle_cat_totals[ck_s][cat] += amt
            cycle_cat_counts[ck_s][cat] += 1
            if cat == "Finance Charge":
                cycle_fc[ck_s] += amt
            if cat == "Other":
                cycle_merch[ck_s][tx][bank] += amt

            row_out = dict(row)
            row_out["category"] = cat
            all_tx.append(row_out)

    all_tx.sort(key=lambda r: (r["statement_date"], r["date"]), reverse=True)

    sorted_cycles = sorted(all_cycles_set)
    latest_cycle  = sorted_cycles[-1] if sorted_cycles else None

    cycles_data = {}
    for i, ck_s in enumerate(sorted_cycles):
        # Categories for this cycle
        cats_out = {}
        for gname, gcfg in MERCHANT_GROUPS.items():
            total = cycle_cat_totals[ck_s].get(gname, 0.0)
            count = cycle_cat_counts[ck_s].get(gname, 0)
            if count > 0:
                cats_out[gname] = {
                    "icon":  gcfg["icon"],
                    "color": CATEGORY_COLORS.get(gname, "#475569"),
                    "total": round(total, 2),
                    "count": count,
                }
        ot = cycle_cat_totals[ck_s].get("Other", 0.0)
        oc = cycle_cat_counts[ck_s].get("Other", 0)
        if oc > 0:
            cats_out["Other"] = {
                "icon": "\U0001f4e6",
                "color": CATEGORY_COLORS["Other"],
                "total": round(ot, 2),
                "count": oc,
            }

        # Top 3 uncategorised merchants for this cycle
        top3 = []
        for tx_name, bspend in cycle_merch[ck_s].items():
            total    = sum(bspend.values())
            top_bank = max(bspend, key=lambda b: bspend[b])
            top3.append({"name": tx_name, "total": round(total, 2), "bank": top_bank})
        top3.sort(key=lambda x: x["total"], reverse=True)
        top3 = top3[:3]

        # Cycle comparison (vs previous cycle)
        cycle_comp = None
        if i > 0:
            prev_ck_s = sorted_cycles[i - 1]
            curr_t    = round(cycle_total[ck_s], 2)
            prev_t    = round(cycle_total[prev_ck_s], 2)
            pct       = round(((curr_t - prev_t) / abs(prev_t) * 100) if prev_t else 0.0, 1)
            curr_fc   = round(cycle_fc[ck_s], 2)
            prev_fc   = round(cycle_fc[prev_ck_s], 2)
            fc_pct    = round((curr_fc - prev_fc) / prev_fc * 100, 1) if prev_fc > 0 else None
            cycle_comp = {
                "current":             {"ck": ck_s,     "label": cycle_label(ck_s),      "total": curr_t},
                "previous":            {"ck": prev_ck_s, "label": cycle_label(prev_ck_s), "total": prev_t},
                "pct_change":          pct,
                "prev_finance_charge": prev_fc,
                "finance_charge_pct":  fc_pct,
            }

        # Bank statement dates (latest per bank in this cycle, formatted)
        bank_stmt_dates = {}
        for bank in BANK_ORDER:
            d = cycle_bank_stmts[ck_s].get(bank, "")
            bank_stmt_dates[bank] = fmt_stmt_date(d) if d else ""

        cycles_data[ck_s] = {
            "total":           round(cycle_total[ck_s], 2),
            "bank_totals":     {b: round(cycle_bank_totals[ck_s].get(b, 0.0), 2) for b in BANK_ORDER},
            "bank_counts":     {b: cycle_bank_counts[ck_s].get(b, 0)            for b in BANK_ORDER},
            "bank_stmt_dates": bank_stmt_dates,
            "finance_charge":  round(cycle_fc[ck_s], 2),
            "categories":      cats_out,
            "top3":            top3,
            "cycle_comp":      cycle_comp,
        }

    # ── Yearly aggregation ────────────────────────────────────────────────
    all_years_set = set()
    year_totals_d = defaultdict(float)
    year_bank_tot = defaultdict(lambda: defaultdict(float))
    year_bank_cnt = defaultdict(lambda: defaultdict(int))
    year_bank_stm = defaultdict(lambda: defaultdict(str))
    year_fc_d     = defaultdict(float)
    year_cat_tot  = defaultdict(lambda: defaultdict(float))
    year_cat_cnt  = defaultdict(lambda: defaultdict(int))

    for ck_s, cd in cycles_data.items():
        yr = ck_s[:4]
        all_years_set.add(yr)
        year_totals_d[yr] += cd["total"]
        year_fc_d[yr]     += cd["finance_charge"]
        for bank in BANK_ORDER:
            year_bank_tot[yr][bank] += cd["bank_totals"].get(bank, 0.0)
            year_bank_cnt[yr][bank] += cd["bank_counts"].get(bank, 0)
            raw_d = cycle_bank_stmts[ck_s].get(bank, "")
            if raw_d and raw_d > year_bank_stm[yr].get(bank, ""):
                year_bank_stm[yr][bank] = raw_d
        for cat, info in cd["categories"].items():
            year_cat_tot[yr][cat] += info["total"]
            year_cat_cnt[yr][cat] += info["count"]

    all_years_sorted = sorted(all_years_set)
    years_data = {}
    for i, yr in enumerate(all_years_sorted):
        cats_yr = {}
        for gname, gcfg in MERCHANT_GROUPS.items():
            tot = year_cat_tot[yr].get(gname, 0.0)
            cnt = year_cat_cnt[yr].get(gname, 0)
            if cnt > 0:
                cats_yr[gname] = {
                    "icon":  gcfg["icon"],
                    "color": CATEGORY_COLORS.get(gname, "#475569"),
                    "total": round(tot, 2),
                    "count": cnt,
                }
        ot = year_cat_tot[yr].get("Other", 0.0)
        oc = year_cat_cnt[yr].get("Other", 0)
        if oc > 0:
            cats_yr["Other"] = {
                "icon": "\U0001f4e6",
                "color": CATEGORY_COLORS["Other"],
                "total": round(ot, 2),
                "count": oc,
            }
        year_comp = None
        if i > 0:
            prev_yr = all_years_sorted[i - 1]
            curr_t  = round(year_totals_d[yr], 2)
            prev_t  = round(year_totals_d[prev_yr], 2)
            pct     = round(((curr_t - prev_t) / abs(prev_t) * 100) if prev_t else 0.0, 1)
            curr_fc = round(year_fc_d[yr], 2)
            prev_fc = round(year_fc_d[prev_yr], 2)
            fc_pct  = round((curr_fc - prev_fc) / prev_fc * 100, 1) if prev_fc > 0 else None
            year_comp = {
                "current":             {"yr": yr,      "total": curr_t},
                "previous":            {"yr": prev_yr, "total": prev_t},
                "pct_change":          pct,
                "prev_finance_charge": prev_fc,
                "finance_charge_pct":  fc_pct,
            }
        bank_stmt_yr = {b: (fmt_stmt_date(year_bank_stm[yr].get(b, "")) if year_bank_stm[yr].get(b) else "") for b in BANK_ORDER}
        years_data[yr] = {
            "total":           round(year_totals_d[yr], 2),
            "bank_totals":     {b: round(year_bank_tot[yr].get(b, 0.0), 2) for b in BANK_ORDER},
            "bank_counts":     {b: year_bank_cnt[yr].get(b, 0)            for b in BANK_ORDER},
            "bank_stmt_dates": bank_stmt_yr,
            "finance_charge":  round(year_fc_d[yr], 2),
            "categories":      cats_yr,
            "year_comp":       year_comp,
        }

    all_cats_seen = sorted({tx["category"] for tx in all_tx})

    # Spending summary — all-time categories (sum across all years)
    at_cat_tot = defaultdict(float)
    at_cat_cnt = defaultdict(int)
    for yd in years_data.values():
        for cat, info in yd["categories"].items():
            at_cat_tot[cat] += info["total"]
            at_cat_cnt[cat] += info["count"]
    at_cats = {}
    for gname, gcfg in MERCHANT_GROUPS.items():
        tot = at_cat_tot.get(gname, 0.0)
        cnt = at_cat_cnt.get(gname, 0)
        if cnt > 0:
            at_cats[gname] = {
                "icon":  gcfg["icon"],
                "color": CATEGORY_COLORS.get(gname, "#475569"),
                "total": round(tot, 2),
                "count": cnt,
            }
    ot = at_cat_tot.get("Other", 0.0)
    oc = at_cat_cnt.get("Other", 0)
    if oc > 0:
        at_cats["Other"] = {
            "icon": "\U0001f4e6",
            "color": CATEGORY_COLORS["Other"],
            "total": round(ot, 2),
            "count": oc,
        }

    return {
        "all_cycles":    sorted_cycles,
        "latest_cycle":  latest_cycle,
        "cycles_data":   cycles_data,
        "cycle_spend":   {b: {ck: round(v, 2) for ck, v in cycle_spend[b].items()} for b in BANK_ORDER},
        "all_years":     all_years_sorted,
        "latest_year":   all_years_sorted[-1] if all_years_sorted else None,
        "years_data":    years_data,
        "banks":         BANK_ORDER,
        "bank_display":  {b: bank_display_parts(b) for b in BANK_ORDER},
        "bank_colors":   {b: BANK_COLORS.get(b, DEFAULT_BANK_COLOR)             for b in BANK_ORDER},
        "all_tx":        all_tx,
        "stmt_dates":    sorted(stmt_dates_set, reverse=True),
        "all_categories": all_cats_seen,
        "last_updated":  get_last_updated(),
        "spending_summary_all_time": {
            "categories": at_cats,
            "total": round(sum(v["total"] for v in at_cats.values()), 2),
        },
    }


# ── HTML template ─────────────────────────────────────────────────────────────

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>CC Tracker</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&family=Sora:wght@600;700&display=swap" rel="stylesheet">
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4/dist/chart.umd.min.js"></script>
  <style>
    *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
    :root {
      --bg:      #0f172a;
      --card:    #1e293b;
      --border:  #334155;
      --text:    #f1f5f9;
      --muted:   #94a3b8;
      --accent:  #38bdf8;
      --green:   #4ade80;
      --red:     #f87171;
      --hover:   #334155;
    }
    html { scroll-behavior: smooth; }
    body {
      background: var(--bg); color: var(--text);
      font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
      font-size: 14px; line-height: 1.5; min-height: 100vh;
    }
    .container { max-width: 1280px; margin: 0 auto; padding: 28px 20px 64px; }
    .section-title {
      font-size: 0.7rem; font-weight: 600; color: var(--muted);
      text-transform: uppercase; letter-spacing: 0.1em; margin-bottom: 14px;
    }
    .card-shadow { box-shadow: 0 2px 12px rgba(0,0,0,0.4); }

    /* ── Header ── */
    .header { margin-bottom: 28px; }
    .header h1 {
      font-family: 'Sora', sans-serif;
      font-size: clamp(1.7rem, 4vw, 2.3rem);
      font-weight: 700; color: var(--text); letter-spacing: -0.02em;
    }
    .header .subtitle { font-size: 0.78rem; color: var(--muted); margin-top: 5px; }

    /* ── Cycle selector ── */
    .cycle-selector { display: flex; align-items: center; gap: 10px; margin-top: 14px; }
    .cycle-selector label { font-size: 0.78rem; color: var(--muted); white-space: nowrap; }
    .cycle-select {
      background: var(--card); border: 1px solid var(--border); color: var(--text);
      padding: 6px 12px; border-radius: 8px; font-size: 0.8rem; font-family: inherit;
      outline: none; cursor: pointer; transition: border-color 0.15s;
    }
    .cycle-select:focus { border-color: var(--accent); }

    /* ── Summary cards ── */
    .cards-row {
      display: grid; grid-template-columns: repeat(4, 1fr);
      gap: 14px; margin-bottom: 14px;
    }
    .card { border-radius: 12px; padding: 20px 22px; position: relative; overflow: hidden; }
    .card-total { background: var(--bg); border: 1px solid var(--border); }
    .card-lbl {
      font-size: 0.68rem; font-weight: 700; text-transform: uppercase;
      letter-spacing: 0.1em; opacity: 0.75; margin-bottom: 8px;
    }
    .card-line1 { font-family: 'Sora', sans-serif; font-size: 1.05rem; font-weight: 700; line-height: 1.1; }
    .card-line2 {
      font-family: 'Sora', sans-serif; font-size: 0.82rem; font-weight: 600;
      opacity: 0.72; margin-bottom: 10px; min-height: 1.1em;
    }
    .card-value {
      font-size: 1.5rem; font-weight: 700; font-variant-numeric: tabular-nums;
      letter-spacing: -0.02em; line-height: 1.2; margin-bottom: 5px;
    }
    .card-sub  { font-size: 0.73rem; opacity: 0.68; }
    .card-meta { display: block; font-size: 0.68rem; opacity: 0.82; margin-top: 2px; }

    /* ── Category + FC wrapper ── */
    .cat-fc-wrapper {
      display: grid; grid-template-columns: 1fr 252px;
      gap: 14px; margin-bottom: 14px; align-items: stretch;
    }
    .cat-section { background: var(--card); border-radius: 12px; padding: 20px 22px; }
    .cat-layout { display: flex; align-items: center; gap: 32px; }
    .donut-wrap { width: 200px; height: 200px; flex-shrink: 0; }
    .cat-list { flex: 1; min-width: 0; }
    .cat-item {
      display: flex; align-items: center; gap: 10px;
      padding: 7px 0; border-bottom: 1px solid var(--border);
    }
    .cat-item:last-child { border-bottom: none; }
    .cat-dot { width: 10px; height: 10px; border-radius: 50%; flex-shrink: 0; }
    .cat-name {
      flex: 1; font-size: 0.82rem; font-weight: 500;
      min-width: 0; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
    }
    .cat-right { display: flex; flex-direction: column; align-items: flex-end; flex-shrink: 0; }
    .cat-amount { font-size: 0.82rem; font-weight: 700; font-variant-numeric: tabular-nums; white-space: nowrap; }
    .cat-txcount { font-size: 0.7rem; color: var(--muted); }

    /* ── Finance Charge stack ── */
    .fc-stack { display: grid; grid-template-rows: 1fr 1fr; gap: 14px; }
    .fc-arrow {
      font-family: 'Sora', sans-serif; font-size: 1.05rem; font-weight: 700;
      font-variant-numeric: tabular-nums; letter-spacing: -0.02em;
      line-height: 1.4; margin: 8px 0 6px; word-break: break-word;
    }

    /* ── Panel ── */
    .panel { background: var(--card); border-radius: 12px; padding: 20px 22px; }
    .top3-scroll { overflow-x: auto; -webkit-overflow-scrolling: touch; }
    .month-block {
      display: flex; justify-content: space-between; align-items: flex-end;
      padding: 12px 0; border-bottom: 1px solid var(--border);
    }
    .month-block:last-of-type { border-bottom: none; }
    .month-lbl { font-size: 0.78rem; color: var(--muted); margin-bottom: 3px; }
    .month-val { font-size: 1.25rem; font-weight: 700; font-variant-numeric: tabular-nums; }
    .month-pct {
      display: inline-block; margin-top: 12px; padding: 4px 14px;
      border-radius: 6px; font-size: 0.95rem; font-weight: 700;
    }
    .pct-up   { color: var(--green); background: rgba(74,222,128,0.1); }
    .pct-down { color: var(--red);   background: rgba(248,113,113,0.1); }
    .no-data-msg { font-size: 0.82rem; color: var(--muted); padding: 8px 0; }

    /* ── Top 3 ── */
    .top3-item {
      display: flex; align-items: center; gap: 10px;
      padding: 10px 0; border-bottom: 1px solid var(--border);
    }
    .top3-item:last-child { border-bottom: none; }
    .top3-rank   { font-size: 0.7rem; color: var(--muted); min-width: 18px; }
    .top3-name   { flex: 1; font-size: 0.82rem; font-weight: 500; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
    .top3-amount { font-size: 0.82rem; font-weight: 700; font-variant-numeric: tabular-nums; white-space: nowrap; }

    /* ── Bar chart ── */
    .chart-wrap { background: var(--card); border-radius: 12px; padding: 20px 22px; margin-bottom: 14px; }
    .chart-scroll { overflow-x: auto; -webkit-overflow-scrolling: touch; }
    #spendChart { height: 280px; display: block; }

    /* ── Table ── */
    .table-section { background: var(--card); border-radius: 12px; padding: 20px 22px; }
    .filters { display: flex; gap: 10px; margin-bottom: 16px; flex-wrap: wrap; }
    .filters select, .filters input[type="text"] {
      background: var(--bg); border: 1px solid var(--border); color: var(--text);
      padding: 7px 12px; border-radius: 8px; font-size: 0.8rem; font-family: inherit;
      outline: none; transition: border-color 0.15s; cursor: pointer;
    }
    .filters select:focus, .filters input:focus { border-color: var(--accent); }
    .filters input[type="text"] { min-width: 180px; }
    .table-scroll { overflow-x: auto; -webkit-overflow-scrolling: touch; }
    table { width: 100%; border-collapse: collapse; font-size: 0.8rem; min-width: 680px; }
    thead th {
      text-align: left; padding: 9px 12px; border-bottom: 1px solid var(--border);
      color: var(--muted); font-weight: 500; white-space: nowrap;
    }
    tbody td { padding: 9px 12px; border-bottom: 1px solid rgba(51,65,85,0.45); vertical-align: middle; }
    tbody tr:last-child td { border-bottom: none; }
    tbody tr:hover td { background: var(--hover); }
    .td-amount { text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }
    .amt-pos { color: var(--text); }
    .amt-neg { color: var(--green); }
    .badge {
      display: inline-block; padding: 3px 10px; border-radius: 20px;
      font-size: 0.68rem; font-weight: 600; color: #fff; white-space: nowrap;
    }

    /* ── Pagination ── */
    .pagination {
      display: flex; align-items: center; justify-content: space-between;
      margin-top: 16px; gap: 12px; flex-wrap: wrap;
    }
    .page-info { font-size: 0.78rem; color: var(--muted); }
    .page-btns { display: flex; gap: 8px; }
    .btn {
      background: var(--bg); border: 1px solid var(--border); color: var(--text);
      padding: 6px 18px; border-radius: 7px; cursor: pointer;
      font-size: 0.78rem; font-family: inherit; transition: background 0.15s;
    }
    .btn:hover:not(:disabled) { background: var(--hover); }
    .btn:disabled { opacity: 0.3; cursor: not-allowed; }

    /* ── Mobile ── */
    @media (max-width: 768px) {
      .container        { padding: 18px 14px 52px; }
      .cards-row        { grid-template-columns: repeat(2, 1fr); }
      .card-value       { font-size: 1.35rem; }
      .cat-layout       { flex-direction: column; align-items: center; }
      .donut-wrap       { width: 180px; height: 180px; }
      .cat-list         { width: 100%; }
      .cat-fc-wrapper   { grid-template-columns: 1fr; }
      .cycle-select     { width: 100%; }
      .top3-inner       { min-width: 420px; }
    }
  </style>
</head>
<body>
<div class="container">

  <!-- ① Header + Cycle Selector -->
  <header class="header">
    <h1>Hello, Lincoln!</h1>
    <p class="subtitle">Last updated: <span id="lastUpdated"></span></p>
    <div class="cycle-selector">
      <label for="cycleDropdown">Cycle:</label>
      <select class="cycle-select" id="cycleDropdown" onchange="switchCycle(this.value)"></select>
    </div>
  </header>

  <!-- ② Summary Cards -->
  <div class="cards-row" id="summaryCards"></div>

  <!-- ③ Spending by Category + Finance Charge -->
  <div class="cat-fc-wrapper">
    <div class="cat-section card-shadow">
      <div class="section-title">Spending by Category</div>
      <div class="cat-layout">
        <div class="donut-wrap"><canvas id="categoryChart"></canvas></div>
        <div class="cat-list" id="categoryList"></div>
      </div>
    </div>
    <div class="fc-stack">
      <div class="panel card-shadow" id="fcChangeCard"></div>
      <div class="panel card-shadow" id="cycleCompCard"></div>
    </div>
  </div>

  <!-- ④ Top 3 Merchants -->
  <div id="top3Panel" class="panel card-shadow" style="margin-bottom:14px">
    <div class="section-title">Top 3 Merchants</div>
    <div class="top3-scroll">
      <div class="top3-inner" id="top3List"></div>
    </div>
  </div>

  <!-- ⑤ Stacked Bar Chart -->
  <div class="chart-wrap card-shadow">
    <div class="section-title">Spend by Billing Cycle</div>
    <div class="chart-scroll">
      <canvas id="spendChart"></canvas>
    </div>
    <div id="cycleAvg" style="margin-top:12px;font-size:0.78rem;color:var(--muted);text-align:center;"></div>
  </div>

  <!-- ⑥ Spending Summary -->
  <div class="chart-wrap card-shadow" style="margin-bottom:14px">
    <div style="display:flex;align-items:center;gap:12px;margin-bottom:14px">
      <div class="section-title" style="margin-bottom:0">Spending Summary</div>
      <select class="cycle-select" id="summaryYearDropdown" onchange="renderSpendingSummary(this.value)" style="padding:4px 12px;font-size:0.75rem;border-radius:20px"></select>
    </div>
    <div class="table-scroll">
      <table style="min-width:360px">
        <thead>
          <tr>
            <th>Category</th>
            <th style="text-align:right">Total</th>
            <th style="text-align:right">% vs Last Year</th>
          </tr>
        </thead>
        <tbody id="summaryTableBody"></tbody>
      </table>
    </div>
  </div>

  <!-- ⑦ All Transactions -->
  <div class="table-section card-shadow">
    <div class="section-title" style="margin-bottom:16px">All Transactions</div>
    <div class="filters">
      <select id="filterBank" onchange="applyFilters()">
        <option value="">All Banks</option>
      </select>
      <select id="filterStmt" onchange="applyFilters()">
        <option value="">All Statements</option>
      </select>
      <select id="filterCategory" onchange="applyFilters()">
        <option value="">All Categories</option>
      </select>
      <input id="filterSearch" type="text" placeholder="Search merchant&hellip;" oninput="applyFilters()">
    </div>
    <div class="table-scroll">
      <table>
        <thead>
          <tr>
            <th>Statement Date</th><th>Date</th><th>Transaction</th>
            <th style="text-align:right">Amount</th><th>Bank</th>
          </tr>
        </thead>
        <tbody id="txBody"></tbody>
      </table>
    </div>
    <div class="pagination">
      <span id="pageInfo" class="page-info"></span>
      <div class="page-btns">
        <button class="btn" id="btnPrev" onclick="prevPage()">&larr; Prev</button>
        <button class="btn" id="btnNext" onclick="nextPage()">Next &rarr;</button>
      </div>
    </div>
  </div>

</div>
<script>
const D = __DASHBOARD_DATA_JSON__;

var selectedCycle = D.latest_cycle;
var categoryChart = null;

function esc(s) {
  return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}
function fmt(n) {
  return '₱ ' + Number(n).toLocaleString('en-PH', {minimumFractionDigits:2, maximumFractionDigits:2});
}
function fmtCycle(ck) {
  var mn = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
  var y = parseInt(ck.slice(0,4),10), m = parseInt(ck.slice(5,7),10);
  var nm = m === 12 ? 1 : m + 1;
  var ny = m === 12 ? y + 1 : y;
  return mn[m-1] + ' 15 – ' + mn[nm-1] + ' 5 \'' + String(ny).slice(-2);
}
function bankColor(b) { return (D.bank_colors && D.bank_colors[b]) || '#6b7280'; }
function fmtShortDate(s) {
  if (!s) return '—';
  var mn = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];
  var parts = s.split(' ')[0].split('-');
  if (parts.length < 3) return s;
  return mn[parseInt(parts[1],10)-1] + ' ' + parseInt(parts[2],10);
}

/* ① Header */
document.getElementById('lastUpdated').textContent = D.last_updated || '';

/* Cycle dropdown */
(function() {
  if (!D.all_cycles || !D.all_cycles.length) return;
  var sel = document.getElementById('cycleDropdown');
  D.all_cycles.slice().reverse().forEach(function(ck) {
    var opt = document.createElement('option');
    opt.value = ck;
    opt.textContent = fmtCycle(ck);
    if (ck === selectedCycle) opt.selected = true;
    sel.appendChild(opt);
  });
})();

function switchCycle(ck) {
  selectedCycle = ck;
  document.getElementById('cycleDropdown').value = ck;
  document.getElementById('top3Panel').style.display = '';
  var cd = D.cycles_data && D.cycles_data[ck];
  if (!cd) return;
  renderSummaryCards(cd, ck);
  renderCategorySection(cd);
  renderFCCards(cd);
  renderCycleCompCard(cd);
  renderTop3(cd);
}

/* ② Summary Cards */
function renderSummaryCards(cd, ck) {
  var wrap = document.getElementById('summaryCards');
  wrap.innerHTML = '';
  var sublabel = fmtCycle(ck);

  var tot = document.createElement('div');
  tot.className = 'card card-total card-shadow';
  tot.innerHTML =
    '<div class="card-lbl">Total</div>' +
    '<div class="card-value">' + fmt(cd.total) + '</div>' +
    '<div class="card-sub">' + esc(sublabel) + '</div>';
  wrap.appendChild(tot);

  D.banks.forEach(function(bank) {
    var parts = D.bank_display[bank] || [bank, ''];
    var count = (cd.bank_counts && cd.bank_counts[bank]) || 0;
    var total = (cd.bank_totals && cd.bank_totals[bank]) || 0;
    var stmtDate = (cd.bank_stmt_dates && cd.bank_stmt_dates[bank]) || '';
    var c = document.createElement('div');
    c.className = 'card card-shadow';
    c.style.background = bankColor(bank);
    c.innerHTML =
      '<div class="card-line1">' + esc(parts[0]) + '</div>' +
      '<div class="card-line2">' + esc(parts[1]) + '</div>' +
      '<div class="card-value">' + fmt(total) + '</div>' +
      '<div class="card-sub">' +
        count.toLocaleString() + ' transaction' + (count !== 1 ? 's' : '') +
        (stmtDate ? '<span class="card-meta">Statement: ' + esc(stmtDate) + '</span>' : '') +
        '<span class="card-meta">Last fetched: ' + esc(fmtShortDate(D.last_updated)) + '</span>' +
      '</div>';
    wrap.appendChild(c);
  });
}

/* ③ Category Donut + List */
function renderCategorySection(cd) {
  var entries = Object.entries(cd.categories || {}).sort(function(a,b){ return b[1].total - a[1].total; });
  var labels  = entries.map(function(e){ return e[0]; });
  var values  = entries.map(function(e){ return e[1].total; });
  var colors  = entries.map(function(e){ return e[1].color; });

  if (categoryChart) {
    categoryChart.data.labels = labels;
    categoryChart.data.datasets[0].data = values;
    categoryChart.data.datasets[0].backgroundColor = colors;
    categoryChart.update();
  } else if (entries.length) {
    categoryChart = new Chart(document.getElementById('categoryChart'), {
      type: 'doughnut',
      data: { labels: labels, datasets: [{ data: values, backgroundColor: colors, borderWidth: 0, hoverOffset: 8 }] },
      options: {
        responsive: true, maintainAspectRatio: false, cutout: '65%',
        plugins: {
          legend: { display: false },
          tooltip: { callbacks: { label: function(ctx){ return '  ' + ctx.label + ': ' + fmt(ctx.raw); } } }
        }
      }
    });
  }

  var listEl = document.getElementById('categoryList');
  if (!entries.length) {
    listEl.innerHTML = '<p class="no-data-msg">No category data for this cycle.</p>'; return;
  }
  listEl.innerHTML = entries.map(function(e) {
    var name = e[0], info = e[1];
    return '<div class="cat-item">' +
      '<span class="cat-dot" style="background:' + info.color + '"></span>' +
      '<span class="cat-name">' + esc(info.icon + ' ' + name) + '</span>' +
      '<div class="cat-right">' +
        '<span class="cat-amount">' + fmt(info.total) + '</span>' +
        '<span class="cat-txcount">' + info.count + ' txn' + (info.count !== 1 ? 's' : '') + '</span>' +
      '</div></div>';
  }).join('');
}

/* Finance Charge card */
function renderFCCards(cd) {
  var fcChEl = document.getElementById('fcChangeCard');
  var fc     = cd.finance_charge || 0;
  var comp   = cd.cycle_comp;
  if (comp && comp.finance_charge_pct !== null && comp.finance_charge_pct !== undefined) {
    var pct = comp.finance_charge_pct, up = pct >= 0;
    fcChEl.innerHTML =
      '<div class="section-title">Finance Charge vs Prev Cycle</div>' +
      '<div class="fc-arrow">' + fmt(comp.prev_finance_charge) + ' → ' + fmt(fc) + '</div>' +
      '<span class="month-pct ' + (up ? 'pct-up' : 'pct-down') + '" style="font-size:0.85rem;padding:3px 10px;">' +
        (up ? '+' : '') + pct + '%' +
      '</span>';
  } else if (comp) {
    fcChEl.innerHTML =
      '<div class="section-title">Finance Charge vs Prev Cycle</div>' +
      '<p class="no-data-msg" style="margin-top:8px">No FC in previous cycle.</p>';
  } else {
    fcChEl.innerHTML =
      '<div class="section-title">Finance Charge vs Prev Cycle</div>' +
      '<p class="no-data-msg" style="margin-top:8px">First cycle on record.</p>';
  }
}

/* FC stack — Prev Cycle vs Current Cycle card */
function renderCycleCompCard(cd) {
  var el = document.getElementById('cycleCompCard');
  el.innerHTML = '<div class="section-title">Prev Cycle Total vs Current Cycle Total</div>';
  if (!cd.cycle_comp) {
    el.innerHTML += '<p class="no-data-msg" style="margin-top:8px">Not enough data yet.</p>'; return;
  }
  var c = cd.cycle_comp, up = c.pct_change >= 0;
  el.innerHTML +=
    '<div class="month-block"><div>' +
      '<div class="month-lbl" style="font-size:0.72rem">' + esc(c.previous.label) + '</div>' +
      '<div class="month-val" style="font-size:0.9rem">' + fmt(c.previous.total) + '</div>' +
    '</div></div>' +
    '<div class="month-block"><div>' +
      '<div class="month-lbl" style="font-size:0.72rem">' + esc(c.current.label) + '</div>' +
      '<div class="month-val" style="font-size:0.9rem">' + fmt(c.current.total) + '</div>' +
    '</div></div>' +
    '<span class="month-pct ' + (up ? 'pct-up' : 'pct-down') + '" style="font-size:0.82rem;padding:3px 10px;margin-top:8px;">' +
      (up ? '+' : '') + c.pct_change + '%</span>';
}

/* ④ Top 3 Merchants */
function renderTop3(cd) {
  var el = document.getElementById('top3List');
  if (!cd.top3 || !cd.top3.length) {
    el.innerHTML = '<p class="no-data-msg">No uncategorized merchants in this cycle.</p>'; return;
  }
  el.innerHTML = cd.top3.map(function(m, i) {
    var color = bankColor(m.bank);
    return '<div class="top3-item">' +
      '<span class="top3-rank">#' + (i+1) + '</span>' +
      '<span class="top3-name" title="' + esc(m.name) + '">' + esc(m.name) + '</span>' +
      '<span class="top3-amount">' + fmt(m.total) + '</span>' +
      '<span class="badge" style="background:' + color + '">' + esc(m.bank.replace(/_/g,' ')) + '</span>' +
    '</div>';
  }).join('');
}

/* Initialize with latest cycle */
if (D.latest_cycle && D.cycles_data && D.cycles_data[D.latest_cycle]) {
  switchCycle(D.latest_cycle);
}

/* ⑤ Bar Chart — always shows all cycles, never changes with cycle selector */
(function() {
  var cycles = D.all_cycles;
  var canvas = document.getElementById('spendChart');
  if (!cycles || !cycles.length) {
    canvas.parentElement.innerHTML = '<p class="no-data-msg" style="padding:24px 0">No transaction data yet.</p>';
    return;
  }
  canvas.style.minWidth = Math.max(480, cycles.length * 140) + 'px';
  var cycleSum = cycles.reduce(function(s,ck){ return s + (D.cycles_data[ck] ? D.cycles_data[ck].total : 0); }, 0);
  document.getElementById('cycleAvg').textContent = 'Avg per cycle: ' + fmt(cycleSum / cycles.length);
  new Chart(canvas, {
    type: 'bar',
    data: {
      labels: cycles.map(fmtCycle),
      datasets: D.banks.map(function(bank) {
        return {
          label: bank.replace(/_/g,' '),
          data:  cycles.map(function(c){ return (D.cycle_spend[bank] && D.cycle_spend[bank][c]) || 0; }),
          backgroundColor: bankColor(bank),
          borderRadius: 3, borderSkipped: false,
        };
      })
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      interaction: { mode: 'index', intersect: false },
      plugins: {
        legend: { position:'bottom', labels:{color:'#94a3b8', font:{family:'Inter',size:12}, padding:20} },
        tooltip: { callbacks: {
          label:  function(ctx){ return '  ' + ctx.dataset.label + ': ' + fmt(ctx.raw); },
          footer: function(items){ var t=items.reduce(function(s,i){return s+i.raw;},0); return 'Total: '+fmt(t); }
        }}
      },
      scales: {
        x: { stacked:true, ticks:{color:'#94a3b8',font:{family:'Inter',size:11}}, grid:{color:'rgba(51,65,85,0.4)'} },
        y: { stacked:true,
          ticks:{ color:'#94a3b8', font:{family:'Inter',size:11},
            callback: function(v){ return '₱' + v.toLocaleString('en-PH'); }
          },
          grid:{color:'rgba(51,65,85,0.4)'}
        }
      }
    }
  });
})();

/* ⑥ Spending Summary */
(function() {
  var sel = document.getElementById('summaryYearDropdown');
  var optAll = document.createElement('option');
  optAll.value = 'all'; optAll.textContent = 'Total All Time';
  sel.appendChild(optAll);
  (D.all_years || []).slice().reverse().forEach(function(yr) {
    var opt = document.createElement('option');
    opt.value = yr; opt.textContent = yr;
    sel.appendChild(opt);
  });
  var initial = D.latest_year || 'all';
  sel.value = initial;
  renderSpendingSummary(initial);
})();

function renderSpendingSummary(mode) {
  var cats, prevCats = null;
  if (mode === 'all') {
    var st = D.spending_summary_all_time || {};
    cats = st.categories || {};
  } else {
    var yd = D.years_data && D.years_data[mode];
    cats = yd ? (yd.categories || {}) : {};
    var idx = (D.all_years || []).indexOf(mode);
    if (idx > 0) {
      var prevYr = D.all_years[idx - 1];
      var pyd = D.years_data && D.years_data[prevYr];
      prevCats = pyd ? (pyd.categories || {}) : null;
    }
  }
  var entries = Object.entries(cats).sort(function(a,b){ return b[1].total - a[1].total; });
  var grandTotal = entries.reduce(function(s,e){ return s + e[1].total; }, 0);
  var rows = entries.map(function(e) {
    var cat = e[0], info = e[1];
    var pctCell = '<td style="text-align:right;color:var(--muted)">—</td>';
    if (prevCats && prevCats[cat] && prevCats[cat].total > 0) {
      var pct = (info.total - prevCats[cat].total) / prevCats[cat].total * 100;
      var up = pct >= 0;
      pctCell = '<td style="text-align:right"><span class="month-pct ' + (up ? 'pct-up' : 'pct-down') +
        '" style="font-size:0.75rem;padding:2px 9px;margin-top:0">' +
        (up ? '+' : '') + pct.toFixed(1) + '%</span></td>';
    }
    return '<tr>' +
      '<td style="display:flex;align-items:center;gap:8px;min-height:36px">' +
        '<span style="width:8px;height:8px;border-radius:50%;background:' + info.color +
          ';flex-shrink:0;display:inline-block"></span>' +
        esc(info.icon + ' ' + cat) +
      '</td>' +
      '<td class="td-amount">' + fmt(info.total) + '</td>' +
      pctCell +
    '</tr>';
  });
  rows.push('<tr style="font-weight:700;border-top:2px solid var(--border)">' +
    '<td style="color:var(--muted);font-size:0.75rem;text-transform:uppercase;letter-spacing:0.08em;padding-top:11px">Total</td>' +
    '<td class="td-amount" style="padding-top:11px">' + fmt(grandTotal) + '</td>' +
    '<td></td></tr>');
  document.getElementById('summaryTableBody').innerHTML = rows.join('');
}

/* ⑦ Transaction Table — always shows all transactions, unchanged */
var filteredTx  = [];
var currentPage = 1;
var PAGE_SIZE   = 20;

(function initTable() {
  var bSel = document.getElementById('filterBank');
  D.banks.forEach(function(b) {
    var o = document.createElement('option');
    o.value = b; o.textContent = b.replace(/_/g,' ');
    bSel.appendChild(o);
  });

  var sSel = document.getElementById('filterStmt');
  (D.stmt_dates || []).forEach(function(d) {
    var o = document.createElement('option');
    o.value = o.textContent = d;
    sSel.appendChild(o);
  });

  var cSel = document.getElementById('filterCategory');
  (D.all_categories || []).forEach(function(cat) {
    var o = document.createElement('option');
    o.value = o.textContent = cat;
    cSel.appendChild(o);
  });

  filteredTx = (D.all_tx || []).slice();
  renderTable();
})();

function applyFilters() {
  var bank  = document.getElementById('filterBank').value;
  var stmt  = document.getElementById('filterStmt').value;
  var cat   = document.getElementById('filterCategory').value;
  var query = document.getElementById('filterSearch').value.toLowerCase();
  filteredTx = (D.all_tx || []).filter(function(tx) {
    return (!bank  || tx.bank === bank) &&
           (!stmt  || tx.statement_date === stmt) &&
           (!cat   || tx.category === cat) &&
           (!query || tx.transaction.toLowerCase().indexOf(query) !== -1);
  });
  currentPage = 1;
  renderTable();
}

function renderTable() {
  var start = (currentPage - 1) * PAGE_SIZE;
  var page  = filteredTx.slice(start, start + PAGE_SIZE);
  document.getElementById('txBody').innerHTML = page.map(function(tx) {
    var neg = tx.amount < 0, color = bankColor(tx.bank);
    return '<tr>' +
      '<td>' + esc(tx.statement_date || '—') + '</td>' +
      '<td>' + esc(tx.date) + '</td>' +
      '<td>' + esc(tx.transaction) + '</td>' +
      '<td class="td-amount ' + (neg ? 'amt-neg' : 'amt-pos') + '">' + fmt(tx.amount) + '</td>' +
      '<td><span class="badge" style="background:' + color + '">' + esc(tx.bank.replace(/_/g,' ')) + '</span></td>' +
    '</tr>';
  }).join('');
  renderPagination();
}

function renderPagination() {
  var total = filteredTx.length;
  var start = total ? (currentPage - 1) * PAGE_SIZE + 1 : 0;
  var end   = Math.min(currentPage * PAGE_SIZE, total);
  document.getElementById('pageInfo').textContent =
    total === 0 ? 'No transactions found' : 'Showing ' + start + '–' + end + ' of ' + total;
  document.getElementById('btnPrev').disabled = currentPage <= 1;
  document.getElementById('btnNext').disabled = end >= total;
}

function prevPage() { if (currentPage > 1) { currentPage--; renderTable(); } }
function nextPage() { if (currentPage * PAGE_SIZE < filteredTx.length) { currentPage++; renderTable(); } }
</script>
</body>
</html>"""


# ── Output & git ──────────────────────────────────────────────────────────────

def generate_html(summary: dict) -> str:
    payload = json.dumps(summary, ensure_ascii=False, separators=(",", ":"))
    return HTML_TEMPLATE.replace("__DASHBOARD_DATA_JSON__", payload)


def write_output(html: str):
    DASHBOARD_PATH.parent.mkdir(parents=True, exist_ok=True)
    DOCS_PATH.parent.mkdir(parents=True, exist_ok=True)
    DASHBOARD_PATH.write_text(html, encoding="utf-8")
    DOCS_PATH.write_text(html, encoding="utf-8")
    print(f"[OK]   Dashboard written → {DASHBOARD_PATH}")
    print(f"[OK]   Dashboard copied  → {DOCS_PATH}")


def git_push(timestamp: str):
    try:
        subprocess.run(["git", "add", str(DASHBOARD_PATH), str(DOCS_PATH)], check=True)
        subprocess.run(["git", "commit", "-m", f"Dashboard updated {timestamp}"], check=True)
        subprocess.run(["git", "push"], check=True)
        print("[OK]   Pushed to GitHub Pages")
    except subprocess.CalledProcessError as e:
        print(f"[WARN] Git push failed: {e} — dashboard written locally but not published")


def run():
    data    = load_transactions()
    summary = compute_summaries(data)
    html    = generate_html(summary)
    write_output(html)
    git_push(summary["last_updated"])
    return summary


if __name__ == "__main__":
    summary = run()
    active  = sum(1 for b in BANK_ORDER if any(
        cd.get("bank_counts", {}).get(b, 0) > 0
        for cd in summary["cycles_data"].values()
    ))
    print(f"\nDone. {active} active bank(s), {len(summary['all_tx'])} total transaction(s).")
    print(f"  Cycles: {summary['all_cycles']}")
    print(f"  Latest: {summary['latest_cycle']}")
    print(f"  Local : {DASHBOARD_PATH}")
    print(f"  Pages : https://lynx-build.github.io/cc-tracker/")
