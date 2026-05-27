import argparse
import hashlib
import json
import re
from pathlib import Path

import openpyxl
from openpyxl import Workbook

PARSED_DIR = Path("parsed")
EXCEL_PATH = Path("data/cc_transactions.xlsx")
HEADERS    = ["Statement Date", "Date", "Transaction", "Amount", "Hash"]

# Column indices (0-based)
COL_STMT_DATE   = 0
COL_DATE        = 1
COL_TRANSACTION = 2
COL_AMOUNT      = 3
COL_HASH        = 4


def compute_hash(bank: str, date: str, transaction: str, amount) -> str:
    raw = f"{bank}{date}{transaction}{amount}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


def get_or_create_sheet(wb: Workbook, bank: str):
    if bank in wb.sheetnames:
        return wb[bank]
    ws = wb.create_sheet(title=bank)
    ws.append(HEADERS)
    return ws


def load_existing_hashes(ws) -> set:
    hashes = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[COL_HASH]:
            hashes.add(row[COL_HASH])
    return hashes


def sort_sheet_descending(ws) -> None:
    """Sort all data rows: statement_date desc, date desc."""
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            data.append(list(row))
    if not data:
        return

    def sort_key(row):
        stmt = row[COL_STMT_DATE] or row[COL_DATE] or ""
        date = row[COL_DATE] or ""
        return (stmt, date)

    data.sort(key=sort_key, reverse=True)

    # Clear existing data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    # Delete the now-empty rows (shrink sheet)
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # Rewrite sorted data
    for row_data in data:
        ws.append(row_data)


def write_bank(wb: Workbook, bank: str, transactions: list) -> tuple[int, int]:
    ws = get_or_create_sheet(wb, bank)
    existing_hashes = load_existing_hashes(ws)

    appended = 0
    skipped  = 0

    for tx in transactions:
        stmt_date   = tx.get("statement_date", "")
        date        = tx.get("date", "")
        transaction = tx.get("transaction", "")
        amount      = tx.get("amount", 0)
        tx_hash     = compute_hash(bank, date, transaction, amount)

        if tx_hash in existing_hashes:
            skipped += 1
            continue

        ws.append([stmt_date, date, transaction, amount, tx_hash])
        existing_hashes.add(tx_hash)
        appended += 1

    if appended:
        sort_sheet_descending(ws)

    return appended, skipped


def dedup_fix(wb: Workbook) -> dict:
    removed = {}
    for bank in wb.sheetnames:
        ws = wb[bank]
        seen = set()
        rows_to_delete = []
        for row in ws.iter_rows(min_row=2):
            tx_hash = row[COL_HASH].value
            if tx_hash in seen:
                rows_to_delete.append(row[0].row)
            elif tx_hash:
                seen.add(tx_hash)
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)
        removed[bank] = len(rows_to_delete)
    return removed


def stmt_date_from_json_name(json_path: Path, bank: str) -> str:
    """Extract approximate statement_date from old-format JSON filenames."""
    stem = json_path.stem  # e.g. BDO_2026-05_12345
    suffix = stem[len(bank) + 1:]  # strip "BDO_" or "BPI_Lincoln_"
    m = re.match(r'^(\d{4}-\d{2}-\d{2})', suffix)
    if m:
        return m.group(1)
    m = re.match(r'^(\d{4}-\d{2})', suffix)
    if m:
        return m.group(1) + "-01"
    return ""


def run(bank_filter=None, fix_dupes=False):
    EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)

    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if fix_dupes:
        removed = dedup_fix(wb)
        wb.save(EXCEL_PATH)
        print("Dedup fix complete:")
        for bank, count in removed.items():
            print(f"  {bank}: {count} duplicate(s) removed")
        return {}

    json_files = sorted(PARSED_DIR.glob("*.json"))
    if not json_files:
        print("[WARN] No parsed JSON files found in parsed/")
        return {}

    # Collect known bank names from statements/ dirs (longest-first to avoid prefix collisions)
    from pathlib import Path as _Path
    _stmts = _Path("statements")
    known_banks = sorted(
        (d.name for d in _stmts.iterdir() if d.is_dir()) if _stmts.exists() else [],
        key=len, reverse=True
    )

    results = {}

    for json_path in json_files:
        stem = json_path.stem  # e.g. "BPI_Lincoln_BPI eStatement-..."
        bank = None
        for b in known_banks:
            if stem.startswith(b + "_"):
                bank = b
                break
        if not bank:
            # Fallback: first underscore-delimited token
            bank = stem.split("_", 1)[0]

        if bank_filter and bank != bank_filter:
            continue

        with open(json_path, encoding="utf-8") as f:
            transactions = json.load(f)

        # Back-fill statement_date for old-format JSON that lacks it
        for tx in transactions:
            if not tx.get("statement_date"):
                tx["statement_date"] = stmt_date_from_json_name(json_path, bank)

        appended, skipped = write_bank(wb, bank, transactions)
        if bank not in results:
            results[bank] = {"appended": 0, "skipped": 0}
        results[bank]["appended"] += appended
        results[bank]["skipped"]  += skipped
        print(f"[OK]   {bank} ({json_path.name}): {appended} added, {skipped} skipped")

    wb.save(EXCEL_PATH)
    print(f"\nSaved → {EXCEL_PATH}")
    return results


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Phase 3: Write parsed transactions to Excel")
    parser.add_argument("--bank", help="Only write this bank (e.g. BDO)")
    parser.add_argument("--dedup-fix", action="store_true")
    args = parser.parse_args()

    results = run(bank_filter=args.bank, fix_dupes=args.dedup_fix)
    if results:
        total_added   = sum(v["appended"] for v in results.values())
        total_skipped = sum(v["skipped"]  for v in results.values())
        print(f"\nDone. {total_added} row(s) added, {total_skipped} skipped.")
